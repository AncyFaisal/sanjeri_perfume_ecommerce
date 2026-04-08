# views/admin_views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse
import csv
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, Case, When, Value, Max
from django.db.models.functions import Coalesce
from django.contrib import messages
from django.db import models
from ..models import CustomUser  # Import your CustomUser model
from ..forms import UserSearchForm, UserFilterForm
from django.db.models import Count, Sum
from ..models import Coupon, Order,Product,ProductVariant
from decimal import Decimal
from django.utils import timezone
from django.contrib.admin.views.decorators import staff_member_required
from ..models import WalletTransaction, OrderItem
from ..services.wallet_service import WalletService
from django.db.models.functions import TruncMonth, TruncYear, TruncDay
from datetime import datetime, timedelta
import json

def admin_required(function):
    """
    Decorator to ensure user is admin/staff
    """
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, "Please login to access admin panel.")
            return redirect('user_login')  # Redirect to your custom login
        if not request.user.is_staff:
            messages.error(request, "You don't have permission to access this page.")
            return redirect('homepage')  # Redirect to home page
        return function(request, *args, **kwargs)
    return wrapper
@login_required
@admin_required
def toggle_user_status(request, user_id):
    """
    Toggle user status between active and blocked
    """
    print(f"=== TOGGLE USER STATUS CALLED ===")
    print(f"User ID: {user_id}")
    print(f"Method: {request.method}")
    print(f"User: {request.user}")
    print(f"Headers: {dict(request.headers)}")
    
    if request.method == 'POST':
        try:
            user = get_object_or_404(CustomUser, id=user_id, is_staff=False)
            print(f"Found user: {user.username}, Current status: {user.status}")
            
            # Toggle status
            if user.status == 'active':
                user.status = 'blocked'
                user.is_active = False
                action = 'blocked'
                message = f'User {user.get_full_name()} has been blocked.'
            else:
                user.status = 'active'
                user.is_active = True
                action = 'unblocked'
                message = f'User {user.get_full_name()} has been unblocked.'
            
            user.save()
            print(f"User status changed to: {user.status}")
            
            # Return JSON for AJAX, redirect for forms
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'action': action,
                    'new_status': user.status,
                    'message': message
                })
            else:
                messages.success(request, message)
                return redirect('user_list')
                
        except Exception as e:
            error_msg = f'Error toggling user status: {str(e)}'
            print(f"ERROR: {error_msg}")
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg})
            else:
                messages.error(request, error_msg)
                return redirect('user_list')
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
@admin_required
def user_list(request):
    """
    User List Page with search, filter and pagination
    """
    # Get all non-staff users with calculated fields
    # users = CustomUser.objects.filter(is_staff=False).annotate(  # Use CustomUser
    #     total_orders=Count('order', distinct=True),
    #     total_amount_spent=Coalesce(Sum(
    #         Case(
    #             When(order__payment_status='paid', then='order__total_amount'),
    #             default=Value(0),
    #             output_field=models.DecimalField()
    #         )
    #     ), 0),
    #     last_order_date=Max('order__created_at')
    # ).order_by('-date_joined')  # Latest users first

    # Get all non-staff users
    # users = CustomUser.objects.filter(is_staff=False).annotate(
    #     total_orders=Value(0, output_field=IntegerField()),  # Specify output_field
    #     total_amount_spent=Value(0, output_field=DecimalField()),  # Specify output_field
    #     last_order_date=Value(None, output_field=DateTimeField())  # Specify output_field
    # ).order_by('-date_joined')
      # Get all non-staff users without annotations
    users = CustomUser.objects.filter(is_staff=False).order_by('-date_joined')

   # Initialize forms
    search_form = UserSearchForm(request.GET or None)
    filter_form = UserFilterForm(request.GET or None)
    
    # Initialize search_query with empty string
    search_query = ""

   # Search functionality
    if search_form.is_valid():
        search_query = search_form.cleaned_data.get('search_query','')
        if search_query:
            users = users.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(phone__icontains=search_query) |
                Q(username__icontains=search_query)
            )

    # Filter functionality
    if filter_form.is_valid():
        status = filter_form.cleaned_data.get('status')
        registration_date_from = filter_form.cleaned_data.get('registration_date_from')
        registration_date_to = filter_form.cleaned_data.get('registration_date_to')
        total_orders = filter_form.cleaned_data.get('total_orders')
        total_spent = filter_form.cleaned_data.get('total_spent')
        
    # Apply filters
        if status:
            users = users.filter(status=status)
    
        if registration_date_from:
            users = users.filter(date_joined__gte=registration_date_from)
        
        if registration_date_to:
            users = users.filter(date_joined__lte=registration_date_to)
    
    # if total_orders:
    #     if total_orders == '10+':
    #         users = users.filter(total_orders__gte=10)
    #     elif total_orders == '5-10':
    #         users = users.filter(total_orders__range=(5, 10))
    #     elif total_orders == '1-5':
    #         users = users.filter(total_orders__range=(1, 5))
    #     elif total_orders == '0':
    #         users = users.filter(total_orders=0)
    
    # if total_spent:
    #     if total_spent == '5000+':
    #         users = users.filter(total_amount_spent__gte=5000)
    #     elif total_spent == '1000-5000':
    #         users = users.filter(total_amount_spent__range=(1000, 5000))
    #     elif total_spent == '0-1000':
    #         users = users.filter(total_amount_spent__range=(0, 1000))

# Get total count before pagination
    total_users_count = users.count()

    # Pagination
    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_title': 'User Management',
        'users': page_obj,
        'search_query': search_query,
        'total_users': users.count(),
    }
    
    return render(request, 'user_list.html', context)

# sanjeri_app/views/admin_views.py

@login_required
@admin_required
def user_detail(request, user_id):
    """
    User Detail Page with Wallet Transactions
    """
    user = get_object_or_404(CustomUser, id=user_id, is_staff=False)
    
    # Get user's wallet and transactions
    try:
        wallet = user.wallet
        transactions = WalletTransaction.objects.filter(
            wallet=wallet
        ).select_related('order').order_by('-created_at')[:10]  # Last 10 transactions
    except:
        wallet = None
        transactions = []
    
    # Calculate user statistics
    total_orders = 0
    total_amount_spent = 0
    last_order = None

    # Get last order
    # last_order = user_orders.first()
    
    # Calculate average order amount
    if total_orders > 0:
        avg_order_amount = total_amount_spent / total_orders
    else:
        avg_order_amount = 0
        
    context = {
        'page_title': f'User Details - {user.get_full_name()}',
        'user': user,
        'wallet': wallet,
        'transactions': transactions,  # Add transactions to context
        'total_orders': total_orders,
        'total_amount_spent': total_amount_spent,
        'avg_order_amount': avg_order_amount,
        'last_order': None,
    }
    
    return render(request, 'user_detail.html', context)

@login_required
@admin_required
def delete_user(request, user_id):
    """
    Delete user account (soft delete)
    """
    if request.method == 'POST':
        try:
            user = get_object_or_404(CustomUser, id=user_id, is_staff=False)
            user_email = user.email
            
            # Soft delete
            user.is_active = False
            user.status = 'blocked'  # Also block the user
            
            # Handle email - ensure uniqueness
            user.email = f"deleted_{user_id}_{user.email}"
            
            # Handle phone - keep within 15 character limit
            if user.phone:
                # Use shorter prefix to fit within 15 chars
                user.phone = f"del{user_id}"[:15]
            else:
                user.phone = f"del{user_id}"[:15]
            
            user.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'User {user_email} deleted successfully'
                })
            
            messages.success(request, f'User {user_email} deleted successfully')
            return redirect('user_list')
            
        except Exception as e:
            error_msg = f'Error deleting user: {str(e)}'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg})
            
            messages.error(request, error_msg)
            return redirect('user_list')
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

# sanjeri_app/views/admin_views.py

@login_required
@admin_required
def admin_dashboard(request):
    # Your existing code
    total_users = CustomUser.objects.count()
    total_products = Product.objects.filter(is_deleted=False).count()
    total_orders = Order.objects.count()
    total_variants = ProductVariant.objects.filter(is_deleted=False).count()
    total_coupons = Coupon.objects.filter(active=True).count()
    
    # Calculate total revenue from delivered/completed orders
    total_revenue = Order.objects.filter(
        status__in=['delivered', 'completed', 'confirmed']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # ===== FIXED: Chart data based on filter =====
    chart_period = request.GET.get('period', 'monthly')
    
    if chart_period == 'yearly':
        orders_chart = Order.objects.filter(
            status__in=['delivered', 'completed', 'confirmed']
        ).annotate(
            period=TruncYear('created_at')
        ).values('period').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('period')
    elif chart_period == 'daily':
        # Last 30 days
        thirty_days_ago = timezone.now() - timedelta(days=30)
        orders_chart = Order.objects.filter(
            status__in=['delivered', 'completed', 'confirmed'],
            created_at__gte=thirty_days_ago
        ).annotate(
            period=TruncDay('created_at')
        ).values('period').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('period')
    else:  # monthly (default)
        # Last 12 months
        twelve_months_ago = timezone.now() - timedelta(days=365)
        orders_chart = Order.objects.filter(
            status__in=['delivered', 'completed', 'confirmed'],
            created_at__gte=twelve_months_ago
        ).annotate(
            period=TruncMonth('created_at')
        ).values('period').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('period')[:12]
    
    # Prepare chart data for JSON
    chart_labels = []
    chart_revenue = []
    chart_orders = []
    
    for item in orders_chart:
        if chart_period == 'yearly':
            chart_labels.append(item['period'].strftime('%Y'))
        elif chart_period == 'daily':
            chart_labels.append(item['period'].strftime('%d %b'))
        else:
            chart_labels.append(item['period'].strftime('%b %Y'))
        
        chart_revenue.append(float(item['total']))
        chart_orders.append(item['count'])
    
    # ===== FIXED: Best selling products (top 10) =====
    # Using correct field names: unit_price instead of price_at_time
    best_products = OrderItem.objects.filter(
        order__status__in=['delivered', 'completed', 'confirmed']
    ).values(
        'variant__product__id', 
        'variant__product__name',
        'variant__product__slug'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('total_price')  # Using total_price instead of price_at_time
    ).order_by('-total_quantity')[:10]
    
    # ===== FIXED: Best selling categories (top 10) =====
    best_categories = OrderItem.objects.filter(
        order__status__in=['delivered', 'completed', 'confirmed']
    ).values(
        'variant__product__category__id',
        'variant__product__category__name'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('total_price')  # Using total_price
    ).order_by('-total_quantity')[:10]
    
    # ===== FIXED: Best selling brands (top 10) =====
    best_brands = OrderItem.objects.filter(
        order__status__in=['delivered', 'completed', 'confirmed']
    ).values(
        'variant__product__brand'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('total_price')  # Using total_price
    ).order_by('-total_quantity')[:10]
    
    # Recent orders for quick view
    recent_orders = Order.objects.select_related('user').order_by('-created_at')[:5]
    
    # Low stock alerts (products with stock < 5)
    low_stock_variants = ProductVariant.objects.filter(
        stock__lt=5, 
        stock__gt=0,
        is_deleted=False
    ).select_related('product')[:5]
    
    out_of_stock = ProductVariant.objects.filter(
        stock=0,
        is_deleted=False
    ).select_related('product').count()
    
    context = {
        # Your existing context variables
        'total_users': total_users,
        'total_products': total_products,
        'total_orders': total_orders,
        'total_variants': total_variants,
        'total_coupons': total_coupons,
        'total_revenue': total_revenue,
        
        # Chart data
        'chart_labels': json.dumps(chart_labels),
        'chart_revenue': json.dumps(chart_revenue),
        'chart_orders': json.dumps(chart_orders),
        'chart_period': chart_period,
        
        # Best selling data - using corrected field names
        'best_products': best_products,
        'best_categories': best_categories,
        'best_brands': best_brands,
        
        # Additional data
        'recent_orders': recent_orders,
        'low_stock_variants': low_stock_variants,
        'out_of_stock_count': out_of_stock,
        
        # Date ranges
        'today': timezone.now(),
        'start_date': (timezone.now() - timedelta(days=30)).date(),
        'end_date': timezone.now().date(),
    }
    
    return render(request, 'admin_dashboard.html', context)


# Add this AJAX endpoint for dynamic chart updates
@login_required
@admin_required
def dashboard_chart_data(request):
    """AJAX endpoint to get chart data based on period"""
    period = request.GET.get('period', 'monthly')
    
    if period == 'yearly':
        orders_chart = Order.objects.filter(
            status__in=['delivered', 'completed']
        ).annotate(
            period=TruncYear('created_at')
        ).values('period').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('period')
    elif period == 'daily':
        thirty_days_ago = timezone.now() - timedelta(days=30)
        orders_chart = Order.objects.filter(
            status__in=['delivered', 'completed'],
            created_at__gte=thirty_days_ago
        ).annotate(
            period=TruncDay('created_at')
        ).values('period').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('period')
    else:
        twelve_months_ago = timezone.now() - timedelta(days=365)
        orders_chart = Order.objects.filter(
            status__in=['delivered', 'completed'],
            created_at__gte=twelve_months_ago
        ).annotate(
            period=TruncMonth('created_at')
        ).values('period').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('period')[:12]
    
    chart_labels = []
    chart_revenue = []
    chart_orders = []
    
    for item in orders_chart:
        if period == 'yearly':
            chart_labels.append(item['period'].strftime('%Y'))
        elif period == 'daily':
            chart_labels.append(item['period'].strftime('%d %b'))
        else:
            chart_labels.append(item['period'].strftime('%b %Y'))
        
        chart_revenue.append(float(item['total']))
        chart_orders.append(item['count'])
    
    return JsonResponse({
        'labels': chart_labels,
        'revenue': chart_revenue,
        'orders': chart_orders
    })

@login_required
@admin_required
def ledger_book_view(request):
    """
    View to display ledger book filter options
    """
    today = timezone.now().date()
    thirty_days_ago = today - timedelta(days=30)
    
    # Get summary stats for the default period
    orders_count = Order.objects.filter(
        created_at__date__gte=thirty_days_ago,
        created_at__date__lte=today
    ).count()
    
    wallet_count = WalletTransaction.objects.filter(
        created_at__date__gte=thirty_days_ago,
        created_at__date__lte=today
    ).count()
    
    total_transactions = orders_count + wallet_count
    
    context = {
        'start_date': thirty_days_ago,
        'end_date': today,
        'total_transactions': total_transactions,
        'orders_count': orders_count,
        'wallet_count': wallet_count,
        'title': 'Ledger Book - Financial Report'
    }
    return render(request, 'admin/ledger_book.html', context)


@login_required
@admin_required
def generate_ledger_book(request):
    """
    Generate ledger book (financial transactions log) as Excel/CSV
    """
    # Get date range from request (default to last 30 days)
    end_date = request.GET.get('end_date', timezone.now().date())
    start_date = request.GET.get('start_date', (timezone.now() - timedelta(days=30)).date())
    format_type = request.GET.get('format', 'excel')  # excel or csv
    
    # Convert string dates to date objects
    try:
        if isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except:
        start_date = (timezone.now() - timedelta(days=30)).date()
        end_date = timezone.now().date()
    
    # Get all financial transactions (orders and wallet transactions)
    orders = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        payment_status__in=['completed', 'success']  # Only paid orders
    ).select_related('user').order_by('created_at')
    
    wallet_transactions = WalletTransaction.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        status='COMPLETED'  # Only completed transactions
    ).select_related('wallet__user', 'order').order_by('created_at')
    
    # Prepare ledger entries
    ledger_entries = []
    
    # Add order transactions (as DEBIT - money out from customer)
    for order in orders:
        ledger_entries.append({
            'date': order.created_at,
            'transaction_id': f"ORD-{order.order_number}",
            'description': f"Order #{order.order_number} - {order.user.get_full_name() or order.user.username}",
            'debit': float(order.total_amount),  # Money coming in (debit for company)
            'credit': 0,
            'balance': 0,
            'user': order.user.email,
            'type': 'Sale',
            'payment_method': order.payment_method,
            'order_id': order.id
        })
    
    # Add wallet transactions
    for wt in wallet_transactions:
        if wt.transaction_type in ['REFUND', 'WITHDRAWAL']:
            # Refunds and withdrawals are CREDITS (money going out)
            description = f"Wallet {wt.get_transaction_type_display()}"
            if wt.order:
                description += f" - Order #{wt.order.order_number}"
            if wt.reason:
                description += f" ({wt.reason})"
            
            ledger_entries.append({
                'date': wt.created_at,
                'transaction_id': f"WLT-{wt.id}",
                'description': description,
                'debit': 0,
                'credit': float(wt.amount),  # Money going out
                'balance': 0,
                'user': wt.wallet.user.email,
                'type': wt.get_transaction_type_display(),
                'payment_method': 'wallet',
                'order_id': wt.order.id if wt.order else None
            })
        elif wt.transaction_type == 'DEPOSIT':
            # Deposits are DEBITS (money coming in)
            ledger_entries.append({
                'date': wt.created_at,
                'transaction_id': f"WLT-{wt.id}",
                'description': f"Wallet Deposit - {wt.reason or 'Manual top-up'}",
                'debit': float(wt.amount),  # Money coming in
                'credit': 0,
                'balance': 0,
                'user': wt.wallet.user.email,
                'type': 'Deposit',
                'payment_method': 'wallet',
                'order_id': None
            })
    
    # Sort by date
    ledger_entries.sort(key=lambda x: x['date'])
    
    # Calculate running balance
    running_balance = 0
    for entry in ledger_entries:
        running_balance += entry['debit'] - entry['credit']
        entry['balance'] = running_balance
    
    # Generate file based on format
    if format_type == 'csv':
        return generate_csv_ledger(ledger_entries, start_date, end_date)
    else:
        return generate_excel_ledger(ledger_entries, start_date, end_date)


def generate_csv_ledger(ledger_entries, start_date, end_date):
    """Generate CSV format ledger"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="ledger_book_{start_date}_to_{end_date}.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow(['LEDGER BOOK - {} to {}'.format(start_date, end_date)])
    writer.writerow([])
    writer.writerow(['Date', 'Transaction ID', 'Description', 'User', 'Type', 
                     'Payment Method', 'Debit (₹)', 'Credit (₹)', 'Balance (₹)'])
    
    # Write data
    for entry in ledger_entries:
        writer.writerow([
            entry['date'].strftime('%Y-%m-%d %H:%M'),
            entry['transaction_id'],
            entry['description'],
            entry['user'],
            entry['type'],
            entry.get('payment_method', ''),
            f"{entry['debit']:.2f}",
            f"{entry['credit']:.2f}",
            f"{entry['balance']:.2f}"
        ])
    
    # Write summary
    writer.writerow([])
    writer.writerow(['SUMMARY'])
    total_debit = sum(e['debit'] for e in ledger_entries)
    total_credit = sum(e['credit'] for e in ledger_entries)
    final_balance = ledger_entries[-1]['balance'] if ledger_entries else 0
    
    writer.writerow(['Total Debit:', f"₹{total_debit:.2f}"])
    writer.writerow(['Total Credit:', f"₹{total_credit:.2f}"])
    writer.writerow(['Net Flow:', f"₹{total_debit - total_credit:.2f}"])
    writer.writerow(['Final Balance:', f"₹{final_balance:.2f}"])
    
    return response


def generate_excel_ledger(ledger_entries, start_date, end_date):
    """Generate Excel format ledger with formatting"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ledger_book_{start_date}_to_{end_date}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ledger Book"
    
    # Define styles
    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    debit_font = Font(color="FF0000")  # Red for debits
    credit_font = Font(color="008000")  # Green for credits
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f"LEDGER BOOK - {start_date} to {end_date}"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Date', 'Transaction ID', 'Description', 'User', 'Type', 
               'Payment Method', 'Debit (₹)', 'Credit (₹)', 'Balance (₹)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data
    for row, entry in enumerate(ledger_entries, 4):
        # Date
        ws.cell(row=row, column=1, value=entry['date'].strftime('%Y-%m-%d %H:%M'))
        
        # Transaction ID
        ws.cell(row=row, column=2, value=entry['transaction_id'])
        
        # Description
        ws.cell(row=row, column=3, value=entry['description'])
        
        # User
        ws.cell(row=row, column=4, value=entry['user'])
        
        # Type
        ws.cell(row=row, column=5, value=entry['type'])
        
        # Payment Method
        ws.cell(row=row, column=6, value=entry.get('payment_method', ''))
        
        # Debit
        debit_cell = ws.cell(row=row, column=7, value=entry['debit'] if entry['debit'] > 0 else '')
        if entry['debit'] > 0:
            debit_cell.font = debit_font
            debit_cell.number_format = '#,##0.00'
        
        # Credit
        credit_cell = ws.cell(row=row, column=8, value=entry['credit'] if entry['credit'] > 0 else '')
        if entry['credit'] > 0:
            credit_cell.font = credit_font
            credit_cell.number_format = '#,##0.00'
        
        # Balance
        balance_cell = ws.cell(row=row, column=9, value=entry['balance'])
        balance_cell.number_format = '#,##0.00'
        if entry['balance'] < 0:
            balance_cell.font = debit_font
        
        # Apply border to all cells in this row
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = border
    
    # Summary section
    summary_row = len(ledger_entries) + 5
    ws.cell(row=summary_row, column=5, value="SUMMARY:")
    ws.cell(row=summary_row, column=5).font = Font(bold=True)
    
    total_debit = sum(e['debit'] for e in ledger_entries)
    total_credit = sum(e['credit'] for e in ledger_entries)
    final_balance = ledger_entries[-1]['balance'] if ledger_entries else 0
    
    ws.cell(row=summary_row + 1, column=5, value="Total Debit:")
    debit_total_cell = ws.cell(row=summary_row + 1, column=6, value=total_debit)
    debit_total_cell.font = debit_font
    debit_total_cell.number_format = '#,##0.00'
    
    ws.cell(row=summary_row + 2, column=5, value="Total Credit:")
    credit_total_cell = ws.cell(row=summary_row + 2, column=6, value=total_credit)
    credit_total_cell.font = credit_font
    credit_total_cell.number_format = '#,##0.00'
    
    ws.cell(row=summary_row + 3, column=5, value="Net Flow:")
    net_flow_cell = ws.cell(row=summary_row + 3, column=6, value=total_debit - total_credit)
    net_flow_cell.number_format = '#,##0.00'
    
    ws.cell(row=summary_row + 4, column=5, value="Final Balance:")
    final_balance_cell = ws.cell(row=summary_row + 4, column=6, value=final_balance)
    final_balance_cell.font = Font(bold=True)
    final_balance_cell.number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response


@login_required
@admin_required
def ledger_book_preview(request):
    """AJAX endpoint to preview ledger data"""
    end_date = request.GET.get('end_date', timezone.now().date())
    start_date = request.GET.get('start_date', (timezone.now() - timedelta(days=30)).date())
    
    try:
        if isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except:
        start_date = (timezone.now() - timedelta(days=30)).date()
        end_date = timezone.now().date()
    
    # Get counts
    orders_count = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        payment_status__in=['completed', 'success']
    ).count()
    
    wallet_count = WalletTransaction.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        status='COMPLETED'
    ).count()
    
    # Get revenue
    total_revenue = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        payment_status__in=['completed', 'success']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    total_refunds = WalletTransaction.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        transaction_type='REFUND',
        status='COMPLETED'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    return JsonResponse({
        'success': True,
        'data': {
            'orders_count': orders_count,
            'wallet_count': wallet_count,
            'total_transactions': orders_count + wallet_count,
            'total_revenue': float(total_revenue),
            'total_refunds': float(total_refunds),
            'net_flow': float(total_revenue - total_refunds)
        }
    })